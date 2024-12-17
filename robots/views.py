import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils.dateparse import parse_datetime
from .models import RegisteredModel, Robot
from datetime import datetime
from django.shortcuts import render
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Alignment, Font


@csrf_exempt
def create_robot(request):
    """
    Запись в БД по полученному JSON о созданном роботе
    """
    if request.method == 'POST':
        try:
            # Загружаем JSON-данные из запроса
            data = json.loads(request.body.decode('utf-8'))
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Не корректный формат данных. Необходимо загрузить данные в формате JSON'},
                                status=400)

        # Получаем данные из JSON
        model = data.get('model')
        version = data.get('version')
        created_str = data.get('created')

        # Проверка обязательных полей
        if not all([model, version, created_str]):
            return JsonResponse({'error': 'Отсутствуют обязательные поля: модель, версия, создано'}, status=400)

        # Парсим дату из строки
        created = parse_datetime(created_str)
        if not created:
            return JsonResponse({'error': 'Неверный формат даты. Должен быть указанынй формат - YYYY-MM-DD HH:MM:SS'},
                                status=400)

        # Проверка существования модели и версии в RegisteredModel
        try:
            registered_model = RegisteredModel.objects.get(model_name=model, version=version)
        except RegisteredModel.DoesNotExist:
            return JsonResponse({
                'error': f"Модель '{model}' с версией '{version}' не зарегистрирована."
            }, status=400)

        # Создаём запись в базе данных
        robot = Robot.objects.create(registered_model=registered_model, created=created)

        # Возвращаем успешный ответ с данными
        return JsonResponse({
            'message': 'Робот успешно создан',
            'robot': {
                'serial': str(robot.serial),
                'model': robot.registered_model.model_name,
                'version': robot.registered_model.version,
                'created': robot.created.strftime('%Y-%m-%d %H:%M:%S')
            }
        }, status=201)

    return JsonResponse({'error': 'Разрешен только метод POST.'}, status=405)


def robots_report_page(request):
    """
    Отображение страницы с формой выбора периода для отчёта.
    """
    return render(request, 'reports/robots_report.html')


def generate_report(request):
    """
    Генерация отчёта в эксель по заданному периоду изготовления роботов.
    """
    # Получаем даты из запроса
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if not start_date or not end_date:
        return HttpResponse("Ошибка: Выберите корректный период.", status=400)

    # Преобразуем даты из строки в формат datetime
    try:
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return HttpResponse("Ошибка: Неверный формат даты.", status=400)

    # Фильтруем данные по указанному периоду
    robots = Robot.objects.filter(created__range=(start_date, end_date))

    # Группируем данные по моделям и версиям
    grouped_data = {}
    for robot in robots:
        model = robot.registered_model.model_name
        version = robot.registered_model.version
        if model not in grouped_data:
            grouped_data[model] = {}
        if version not in grouped_data[model]:
            grouped_data[model][version] = {'count': 0}
        grouped_data[model][version]['count'] += 1

    # Генерация Excel-файла
    wb = openpyxl.Workbook()
    for model, versions in grouped_data.items():
        ws = wb.create_sheet(title=model)
        ws.append(["Модель", "Версия", "Количество за выбранный"])
        for version, data in versions.items():
            ws.append([model, version, data['count']])  # Количество за период

        # Стилизация заголовков
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Удаляем автоматически созданный пустой лист
    del wb['Sheet']

    # Возврат файла в виде ответа
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response[
        'Content-Disposition'] = f'attachment; filename=robots_report_{start_date.strftime("%Y-%m-%d")}_to_{end_date.strftime("%Y-%m-%d")}.xlsx'
    wb.save(response)
    return response
